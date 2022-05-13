import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os 


# def transformCSVtoXLSX():
csvToXlsx1 = r'C:\Users\Take4\Desktop\1651676165304.2.csv'
csvToXlsx2 = r'C:\Users\Take4\Desktop\1651676165304.2.xlsx'

transform = pd.read_csv(csvToXlsx1, sep=";", encoding = 'latin-1')

transform.to_excel(csvToXlsx2, index=None)

# transformCSVtoXLSX()

# book = load_workbook (csvToXlsx2)
# ws = book.worksheets[0]
# for cell in ws["A"]:
#     if cell.value is None:
#         print (cell.row)
#         break
# else:
#     print (cell.row + 1)

linha = 2

csvToXlsx4 = r'C:\Users\Take4\Desktop\DatabaseResponsáveis.xlsx'

# linhaDB = 1

data10 = load_workbook (csvToXlsx4)
data20 = data10 ['Planilha1'] 
# valor10 = data20.cell(row=linhaDB, column=1).value

#FUNÇÃO QUE CRIA O BANCO DE DADOS RELACIONADO PRA QUE PESSOA O CASO É DIRECIONADO
#ESTÁ LIGADO AO ARQUIVO LISTADO COMO DATABASEREPONSAVEIS - QUE LISTA OS CASOS PARA OS RESPONSÁVEIS
def criandoDBTipoDeCaso(lista1, colunaExcel):
    linhaDB = 1
    valor10 = data20.cell(row=linhaDB, column=colunaExcel).value
    
    while valor10 is not None:
        
        valor10 = data20.cell(row=linhaDB, column=colunaExcel).value
        lista1.append(valor10)
        linhaDB += 1
        
    remocaoNone1 = len(lista1)
    lista1.remove(lista1[remocaoNone1-1])
    
#FUNÇÃO QUE CRIA O BANCO DE DADOS RELACIONADO PRA QUE PESSOA O CASO É DIRECIONADO
#ESTÁ LIGADO AO ARQUIVO LISTADO COMO DATABASEREPONSAVEIS - QUE LISTA OS LOCAIS DOS CASOS PARA OS RESPONSÁVEIS
# def criandoDBLocalDeCaso(lista1, colunaExcel):
#     linhaDB2 = 1
#     valor10 = data20.cell(row=linhaDB2, column=colunaExcel).value
    
#     while valor10 is not None:
        
#         valor10 = data20.cell(row=linhaDB2, column=colunaExcel).value
#         lista1.append(valor10)
#         linhaDB2 += 1
#     remocaoNone1 = len(lista1)
#     lista1.remove(lista1[remocaoNone1-1])

#FUNÇÃO QUE RETORNA O NÚMERO DA COLUNA CONFORME TEXTO ESPECIFICADO NO INPUT.
def achandoColunaCorreta(cabecalho):
    indexColuna = 1
    valorCabecalho = data2.cell(row=1, column=indexColuna).value
    
    while valorCabecalho != cabecalho:
        valorCabecalho = data2.cell(row=1, column=indexColuna).value
        indexColuna += 1
        
    return (indexColuna-1)

#FUNCAO QUE INSERE O NUMERO DO PROTOCOLO AO ARRAY DO RESPONSAVEL
def LeitorLinhasExcel (bancoDeDadosCasos, bancoDeDadosFluxo, bancoDeDadosLocal, colunaQ, colunaT, colunaAL, responsavel):
    itens = (len(bancoDeDadosCasos))
    index = 1
    itens2 = (len(bancoDeDadosLocal))
    index2 = 0
    itens3 = (len(bancoDeDadosFluxo))
    index3 = 0
    while index < itens:
        if bancoDeDadosCasos[index] in colunaQ:
            while index2 < itens2:
                if bancoDeDadosLocal [index2] in colunaAL:
                    while index3 < itens3:
                        if bancoDeDadosFluxo [index3] in colunaT:
                            responsavel.append(protocolo)
                        index3 += 1
                    # try:
                    #     if bancoDeDadosFluxo[1] in colunaT or bancoDeDadosFluxo[2] in colunaT:
                    #         responsavel.append(protocolo)
                    # except:
                    #     if bancoDeDadosFluxo[0] in colunaT:
                    #         responsavel.append(protocolo)
                index2 += 1
        index += 1


data1 = load_workbook (csvToXlsx2)
data2 = data1 ['Sheet1'] 
valor = data2.cell(row=linha, column=1).value

# --------------------------------------------------------------------------------------

TarcisoSilvaDBCasos = []
criandoDBTipoDeCaso (TarcisoSilvaDBCasos, 1)
TarcisoSilvaDBLocais = []
criandoDBTipoDeCaso(TarcisoSilvaDBLocais, 2)
TarcisoDBFluxo = []
criandoDBTipoDeCaso(TarcisoDBFluxo, 3)
TarcisoSilva = []

RosaneCardosoDBCasos = []
criandoDBTipoDeCaso (RosaneCardosoDBCasos, 4)
RosaneCardosoDBLocais = []
criandoDBTipoDeCaso(RosaneCardosoDBLocais, 5)
RosaneCardosoFluxo = []
criandoDBTipoDeCaso(RosaneCardosoFluxo, 6)
RosaneCardoso = []

RosaneCardoso2DBCasos = []
criandoDBTipoDeCaso (RosaneCardoso2DBCasos, 7)
RosaneCardoso2DBLocais = []
criandoDBTipoDeCaso(RosaneCardoso2DBLocais, 8)
RosaneCardoso2Fluxo = []
criandoDBTipoDeCaso(RosaneCardoso2Fluxo, 9)
RosaneCardoso2 = []

TarcisoSilvaEJoseRibamarDBCasos = []
criandoDBTipoDeCaso (TarcisoSilvaEJoseRibamarDBCasos, 10)
TarcisoSilvaEJoseRibamarDBLocais = []
criandoDBTipoDeCaso(TarcisoSilvaEJoseRibamarDBLocais, 11)
TarcisoSilvaEJoseRibamarDBLocaisFluxo = []
criandoDBTipoDeCaso(TarcisoSilvaEJoseRibamarDBLocaisFluxo, 12)
TarcisoSilvaEJoseRibamar = []

# --------------------------------------------------------------------------------------

colunaProtocolo = achandoColunaCorreta("PROTOCOLO")
colunaTipoIncidente = achandoColunaCorreta("TIPO DE INCIDENTE ALIANT")
colunaFluxoDenuncia = achandoColunaCorreta("FLUXO DA DENÚNCIA")
colunaLocalIncNvl0 = achandoColunaCorreta("LOCAL DO INCIDENTE NIVEL 0")

# while valor is not None:
while linha <= 739:
    
    valor = data2.cell(row=linha, column=1).value
    protocolo = data2.cell(row=linha, column=colunaProtocolo).value
    linhaQ = data2.cell(row=linha, column=colunaTipoIncidente).value
    linhaT = data2.cell(row=linha, column=colunaFluxoDenuncia).value
    linhaAL = data2.cell(row=linha, column=colunaLocalIncNvl0).value
    linha += 1
    
    LeitorLinhasExcel(TarcisoSilvaDBCasos, TarcisoDBFluxo, TarcisoSilvaDBLocais, linhaQ, linhaT, linhaAL, TarcisoSilva)
    LeitorLinhasExcel(RosaneCardosoDBCasos, RosaneCardosoFluxo, RosaneCardosoDBLocais, linhaQ, linhaT, linhaAL, RosaneCardoso)
    LeitorLinhasExcel(RosaneCardoso2DBCasos, RosaneCardoso2Fluxo, RosaneCardoso2DBLocais, linhaQ, linhaT, linhaAL, RosaneCardoso2)
    LeitorLinhasExcel(TarcisoSilvaEJoseRibamarDBCasos, TarcisoSilvaEJoseRibamarDBLocaisFluxo, TarcisoSilvaEJoseRibamarDBLocais, linhaQ, linhaT, linhaAL, TarcisoSilvaEJoseRibamar)
    
print ("Casos de " + str(TarcisoSilvaDBCasos[0]))
print (TarcisoSilva)
print ('\n')
print ("Casos de " + str(RosaneCardosoDBCasos[0]))
print (RosaneCardoso)
print ('\n')
print ("Casos de " + str(RosaneCardoso2DBCasos[0]))
print (RosaneCardoso2)
print ('\n')
print ("Casos de " + str(TarcisoSilvaEJoseRibamarDBCasos[0]))
print (TarcisoSilvaEJoseRibamar)












